from datetime import datetime, time, timedelta
import requests
from bs4 import BeautifulSoup
import asyncio
from entrance_data import api_id, api_hash
from telethon import TelegramClient, events
import os
import openpyxl



unames = []
spisok = []
f_spisok = []
f2_spisok = []
alive_bots = []
id_list = []
cl = []
pop_cl = []
getter_name = []
chats = tuple(id_list)


client = TelegramClient('test_other_variant1', api_id, api_hash,system_version="4.16.30-vxCUSTOM")

def name_parse():
    print("it works")
    path = os.path.abspath("../Боты и скрипты.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb.sheetnames
    urls = ws[0]
    getters = ws[1]
    for i in range(0, wb[urls].max_row):
        for col in wb[urls].iter_cols(1, 3):
            if "https" in str(col[i].value):
                spisok.append(str(col[i].value))
    for i in range(1, wb[getters].max_row):
        for col in wb[getters].iter_cols(2):
            getter_name.append(str(col[i].value))
    for link in spisok:
        resp = requests.get(link)
        soup = BeautifulSoup(resp.text, "lxml")
        name = soup.find(class_="tgme_page_title")
        username = soup.find(class_="tgme_page_extra")
        id_list.append(username.text[4:])
async def check_bot(username: str):

        user = await client.get_input_entity(username)
        await client.send_message(user, '/start')
        respond = False
        async with client.conversation(user) as conv:
            try:
                response = await conv.wait_event(events.NewMessage())
                respond = True
                alive_bots.append(f"@{username} активен")
                print(f"{username} активен")
            except asyncio.TimeoutError:
                alive_bots.append(f"@{username} неактивен")
                print(f"{username} неактивен")
        return respond

async def from_list_to_reply(adresat: list, stroka:str):
    for adress_uname in adresat:
        entity = await client.get_entity(adress_uname)
        await client.send_message(entity = entity, message= stroka)

async def main():
    print('Начинаем проверку ботов...')
    await client.start()
    name_parse()
    for bot_username in id_list:
        await check_bot(bot_username)
    for status in alive_bots:
        await from_list_to_reply(getter_name, status)
    await client.disconnect()

async def time_scheduler(target_time):
    now = datetime.now()
    target_datetime = datetime.combine(now.date(), target_time)
    if now.time() > target_time:
        target_datetime += timedelta(days=1)
    await asyncio.sleep((target_datetime - now).total_seconds())

async def daily_job():
    target_time = time(7, 30)
    while True:
        print("Ожидание следующего запуска в 7:30 утра...")
        await time_scheduler(target_time)
        await main()

        await asyncio.sleep(24*3600 - 1)

if __name__ == "__main__":
    with client:
        client.loop.create_task(daily_job())
        client.loop.run_forever()